#########################################################################
# 作者:李宁（蒙娜丽宁），UnityMarvel创始人
#
# 微信公众号：极客起源
#
# B站：https://space.bilibili.com/477001733
#
# Copyright © 2022 Lining. All rights reserved.
#
# 版本: 2.0
#########################################################################

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = '合并单元格'
# 合并单元格
ws.merge_cells('A1:C2')
ws.cell(1,1).value = '合并后的单元格'

# 取消单元格合并
ws.unmerge_cells('A1:C2')


wb.save('cells.xlsx')
