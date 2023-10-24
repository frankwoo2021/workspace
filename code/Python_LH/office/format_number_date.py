#########################################################################
# 作者:李宁（蒙娜丽宁），UnityMarvel创始人
#
# 微信公众号：极客起源
#
# B站：https://space.bilibili.com/477001733
#
# Copyright © 2022 Lining. All rights reserved.
#
# 版本: 2.0
#########################################################################

from openpyxl import Workbook
import datetime
wb = Workbook()

ws = wb.active

ws.title = '格式化数字和日期'
ws.column_dimensions['A'].width = 50

#ws['A2'].value = '124.5685656556'
ws['A2'].value = 124.5685656556
ws['A2'].number_format ='0.00'  # 保留小数点后两位

# 格式化日期
# 方法1：先用Python格式化完日期，然后再给Excel
ws.column_dimensions['C'].width = 50
ws['C2'].value = datetime.datetime(2020,6,1).strftime('今天是：%m/%d/%Y')
print(type(datetime.datetime(2020,6,1).strftime('%m/%d/%Y')))

# 方法2：在Excel中格式化日期
ws['C3'].value = datetime.datetime(2020,6,1)
ws['C3'].number_format = 'yyyy年m月d日'


wb.save('format.xlsx')
