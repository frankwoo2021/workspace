#########################################################################
# 作者:李宁（蒙娜丽宁），UnityMarvel创始人
#
# 微信公众号：极客起源
#
# B站：https://space.bilibili.com/477001733
#
# Copyright © 2022 Lining. All rights reserved.
#
# 版本: 2.0
#########################################################################

'''
1. 获取一定范围的表格（子表格）
2. 输出子表格中所有的cell
3. 获取特定行和列的cell集合
4. 获取多列和多行


'''

import openpyxl

wb = openpyxl.load_workbook('first.xlsx')
sheet = wb['新的Sheet2']
print('获取子表格')
cell_range = sheet['C5':'E6']
print(type(cell_range))
print(cell_range)

print('输出子表格中所有的cell')
for rows in cell_range:  # 对每一行迭代
    for cell in rows:    # 对每一行中的cell迭代
        print(cell)

print('获取特定行和列的cell集合')


rows = sheet['3:6']
print(rows)
# 先行后列
i = 1
for rows in sheet.iter_rows(min_row=1,max_col=3,max_row=2):
    for cell in rows:
        print(cell)
        cell.value = i
        i += 1
print('--------')
# 先列后行
i = 1
for cols in sheet.iter_cols(min_row=3,max_col=3,max_row=4):
    for cell in cols:
        print(cell)
        cell.value = i
        i += 1
wb.save('first.xlsx')
wb.close()
