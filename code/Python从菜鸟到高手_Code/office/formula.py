#########################################################################
# 作者:李宁（蒙娜丽宁），UnityMarvel创始人
#
# 微信公众号：极客起源
#
# B站：https://space.bilibili.com/477001733
#
# Copyright © 2022 Lining. All rights reserved.
#
# 版本: 2.0
#########################################################################

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.title = '使用公式'
ws['B2'] = '=sum(1,2,3,4,5,6,7,8,9,10)'
ws['B3'] ='=sin(3.14/8)'

ws['A1'] = 10
ws['A2'] = 20
ws['A3'] = 30

ws['C1'] = '=sum(A1:A3)'
wb.save('formula.xlsx')
