#########################################################################
# 作者:李宁（蒙娜丽宁），UnityMarvel创始人
#
# 微信公众号：极客起源
#
# B站：https://space.bilibili.com/477001733
#
# Copyright © 2022 Lining. All rights reserved.
#
# 版本: 2.0
#########################################################################

import openpyxl
wb = openpyxl.load_workbook('first.xlsx')

print(wb.sheetnames,'\n')

sheet = wb['新的Sheet2']
c3 = sheet['C3']
print(type(c3))
# 获取单元格的数据
print(f'C3:{c3.value}')
print(f'D11:{sheet["D11"].value}')
print(f'F7:{sheet["F7"].value}')  # Cell的位置，大小写不敏感

sheet['C7'].value = '你好'

sheet['B10'].value = '=sum(C3,D11)'

# 第2种方法
print(sheet.cell(row = 7,column=3).value)
print(sheet.cell(row = 7,column=3,value = 200).value)
'''
for x in range(1,101):
    for y in range(1,101):
        print(sheet.cell(row = x,column=y).value)
'''
print('-----')
print(sheet.max_column)
print(sheet.max_row)
for x in range(1,sheet.max_row):
    for y in range(1,sheet.max_column):
        cell = sheet.cell(row = x,column=y)
        if cell.value != None:
            print(cell.value)
wb.save('first.xlsx')

wb.close()