#########################################################################
# 作者:李宁（蒙娜丽宁），UnityMarvel创始人
#
# 微信公众号：极客起源
#
# B站：https://space.bilibili.com/477001733
#
# Copyright © 2022 Lining. All rights reserved.
#
# 版本: 2.0
#########################################################################

from openpyxl.styles import *
import openpyxl
wb = openpyxl.load_workbook('table1.xlsx')
ws = wb.active
# A22:C24
blueFont = Font(name='Arial',color='0000FF')
redFill = PatternFill('solid',fgColor='FF0000')

#
for col in range(1,4):
    for row in range(22,25):
        ws.cell(row,col).font = blueFont
        ws.cell(row,col).fill = redFill
wb.save('table1.xlsx')
wb.close()