#########################################################################
# 作者:李宁（蒙娜丽宁），UnityMarvel创始人
#
# 微信公众号：极客起源
#
# B站：https://space.bilibili.com/477001733
#
# Copyright © 2022 Lining. All rights reserved.
#
# 版本: 2.0
#########################################################################


from openpyxl import Workbook

wb = Workbook()  # 在内存中创建了一个Excel文档

# 获取默认的sheet
ws = wb.active

print(ws.title)
ws.title = '我的标题'

# 添加新的sheet
ws1 = wb.create_sheet('新的sheet1')  # 最加一个sheet
ws2 = wb.create_sheet('新的Sheet2',0)  # 在第1个位置插入Sheet
ws3 = wb.create_sheet('新的Sheet3',-1)  # 在倒数第2个位置插入

ws.sheet_properties.tabColor = '1072BA'
ws3.sheet_properties.tabColor = 'FFDD00'

wb.save('first.xlsx')


