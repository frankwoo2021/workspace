#########################################################################
# 作者:李宁（蒙娜丽宁），UnityMarvel创始人
#
# 微信公众号：极客起源
#
# B站：https://space.bilibili.com/477001733
#
# Copyright © 2022 Lining. All rights reserved.
#
# 版本: 2.0
#########################################################################

from openpyxl.styles import *
import openpyxl
wb = openpyxl.load_workbook('table1.xlsx')
ws = wb.active
# 设置文字颜色
redFont = Font(name='Arial',color='FF0000')
ws.cell(13,2).font = redFont

# 设置背景颜色
blueFill = PatternFill('solid',fgColor='0000FF')
ws.cell(13,2).fill = blueFill


wb.save('table1.xlsx')
wb.close()